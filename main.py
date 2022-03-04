import timeit
import openpyxl
import os
import xlrd
import traceback

from tkinter import Tk
from tkinter.filedialog import askopenfilenames
from typing import TYPE_CHECKING, List, Optional, Union

from graph import StudentAveragesGraph
from models import Student, Subject, Summary

DEBUG = True

class ParsingError(Exception):

    def __init__(self, message: str) -> None:
        super().__init__(message)

class ConsoleUtils:
    ERROR = '\033[91m'
    HEADER = '\033[95m'
    WARNING = '\033[93m'
    END = '\033[0m'

    def __init__(self) -> None:
        pass

    def input(self, prompt: str = None) -> str:
        if prompt is None:
            return input()
        return input(self.HEADER + prompt + self.END)

    def print(self, *values):
        print(*values)

    def default(self, *text: str):
        self.print('\n'.join(text))

    def warn(self, *text: str):
        self.print(self.WARNING + '\n'.join(text) + self.END)

    def info(self, *text: str):
        self.print(self.HEADER + '\n'.join(text) + self.END)

    def error(self, *text: str):
        self.print(self.ERROR + '\n'.join(text) + self.END)

    def parse_error(self, file: str, message: str):
        self.error(f"Nevertinamas '{file}': {message}")


cu = ConsoleUtils()

xlrdSheet = xlrd.sheet.Sheet
openpyxlSheet = openpyxl.worksheet._read_only.ReadOnlyWorksheet

class UnifiedSheet:
    """A class which implements a unified Sheet object."""

    def __init__(self, sheet: Union[xlrdSheet, openpyxlSheet]) -> None:
        self._sheet = sheet

    @property
    def xlrd(self) -> bool:
        """Returns true if input sheet instance was of xlrd."""
        return isinstance(self._sheet, xlrdSheet)

    def get_cell(self, column: int, row: int) -> Union[None, str, int, float]:
        """Returns cell value at specified column and row.

        The arguments are flipped than the default implementations."""
        if self.xlrd:
            return self._sheet.cell_value(row - 1, column - 1) or None
        return self._sheet.cell(row, column).value

class SpreadsheetReader:
    """A class which implements a unified Excel spreadsheet reader."""

    if TYPE_CHECKING:
        file_path: str

    def __init__(self, original_path: str) -> None:
        self.file_path = original_path

        if self.has_archive_header:
            self._f = open(self.file_path, "rb")
            doc = openpyxl.load_workbook(self._f, data_only=True, read_only=True)
        else:
            doc = xlrd.open_workbook(self.file_path, ignore_workbook_corruption=True)

        self._doc = doc

        # Initialize a unified sheet object to keep my nerves in check
        if isinstance(doc, xlrd.book.Book):
            self.sheet = UnifiedSheet(doc.sheet_by_index(0))
        else:
            assert isinstance(doc, openpyxl.Workbook)
            self.sheet = UnifiedSheet(doc.worksheets[0])

    @property
    def has_archive_header(self) -> bool:
        """Returns True if file contains an archive header.
        Usually infers that the file is of the new, Open XML variety."""
        with open(self.file_path, "rb") as f:
            return f.read(4) == b'PK\x03\x04'

    def close(self) -> None:
        """Closes the reader."""
        if self.sheet.xlrd:
            self._doc.release_resources()
        else:
            if hasattr(self, "_f"):
                self._f.close()
            self._doc.close()
        del self._doc

class DataParser:

    def __init__(self, file_path: str) -> None:
        self._reader = SpreadsheetReader(file_path)
        self._sheet = self._reader.sheet

        # Obtain term start, end and type
        raw_term_value = self.cell(9, 2)
        # Laikotarpis: 2020-2021m.m.II pusmetis -> 2020-2021m.m.II pusmetis
        raw_term_value = raw_term_value.split(": ")[1].strip() # Split here because the word 'Periodas' could also be used
        term_value = raw_term_value.split("-")

        self.term_start = int(term_value[0])
        self.term_end = int(term_value[1][:4])
        self.type = term_value[1][8:]

        self._average_col = None
        self._last_student_row = None

    def close(self) -> None:
        """Closes the reader. No operations should be performed afterwards."""
        self._reader.close()

    def cell(self, col: int, row: int) -> Optional[Union[str, int, float]]:
        """Boilerplate function for returning value at the specified column and row of the cell."""
        return self._sheet.get_cell(col, row)

    def find_average_column(self) -> int:
        """Returns the column for the average mark column."""
        if self._average_col is None:
            off_col = 2
            while True:
                # Row of 'Pasiekimų lygiai'
                value = self.cell(off_col + 2, 3)
                if value is not None:
                    break
                off_col += 1
            self._average_col = off_col + 1
        return self._average_col

    def find_last_student_row(self) -> int:
        """Returns a row of the last student."""
        if self._last_student_row is None:
            off_row = 13
            while True:
                value = self.cell(1, off_row + 1)
                # Encounter string 'Dalyko vidurkis' in a list of digits
                if isinstance(value, str):
                    break
                off_row += 1
            self._last_student_row = off_row
        return self._last_student_row

    def get_school_name(self) -> str:
        """Returns the name of the school."""
        return self.cell(1, 1)[9:]

    def get_grade_name(self) -> str:
        """Returns the name of the grade."""
        return self.cell(9, 1)[7:]

    def get_student_data(self) -> List[Student]:
        """Returns a list of student objects."""
        students = []
        l_row = self.find_last_student_row()
        for row in range(14, l_row + 1):
            student = Student(
                self.cell(2, row),
                self.get_student_subjects(row - 14),
                self.get_student_average(row - 14)
            )
            students.append(student)
        return students

    def get_student_subjects(self, student_idx: int, ignore_modules: int = False) -> List[Subject]:
        """Returns a list of student's subject objects."""
        subjects = []
        a_col = self.find_average_column()
        for col in range(3, a_col):
            subject = Subject(
                self.cell(col, 4),
                self.cell(col, student_idx + 14)
            )
            if subject.is_module and ignore_modules:
                continue
            subjects.append(subject)
        return subjects

    def get_student_average(self, student_idx: int) -> Optional[float]:
        """Returns student's average mark."""
        col = self.find_average_column()
        value = self.cell(col, student_idx + 14)
        if value == 0:
            return None
        return value

    def create_summary(self) -> Summary:
        """Attempts to create a Summary object.

        May raise an exception."""

        # Check whether Excel file is of correct type
        # TODO: investigate future proofing
        if self.cell(1, 2) != "Ataskaita: Mokinių pasiekimų ir lankomumo suvestinė":
            raise ParsingError("Pateiktas ataskaitos tipas yra netinkamas")

        column = self.find_average_column()
        row = self.find_last_student_row()

        # Check whether group average is a zero, if it is, throw parsing error due to incomplete file
        if self.cell(column, row + 1) == 0:
            raise ParsingError((
                "Trūksta duomenų, suvestinė yra nepilna."
                " Įsitikinkite ar pusmetis/trimestras yra tikrai ir pilnai išvestas!"
            ))

        return Summary(
            parser.get_school_name(),
            parser.get_grade_name(),
            self.type,
            (self.term_start, self.term_end),
            parser.get_student_data()
        )


cu.info("Pasirinkite pusmečių/trimestrų suvestinių Excel failus")

Tk().withdraw() # prevent full GUI from appearing
filenames = askopenfilenames(filetypes=[("Excel suvestinių failai", ".xlsx .xls")])

# Do initial creation of summaries by iterating the submitted files and validating them
summaries: List[Summary] = []
for filename in filenames:
    start_time = timeit.default_timer()
    base_name = os.path.basename(filename)

    try:
        parser = DataParser(filename)
        summary = parser.create_summary()
        parser.close()
    except ParsingError as e:
        cu.parse_error(base_name, str(e))
        continue
    except Exception as e:
        cu.parse_error(base_name, "pateikta ne suvestinė arba netinkamas failas")
        if DEBUG:
            print(e)
            traceback.print_tb(e.__traceback__)
        continue

    if summary.type == "metinis":
        cu.parse_error(base_name, "metinė ataskaita yra nevertinama")
        continue

    if any(s for s in summaries if s.representable_name == summary.representable_name):
        cu.parse_error(base_name, "tokia ataskaita jau vieną kartą buvo pateikta ir perskaityta")
        continue

    summaries.append(summary)
    cu.default(f"'{base_name}' perskaitytas")
    if DEBUG:
        print(f"'{base_name}' skaitymas užtruko {timeit.default_timer() - start_time}s")

if len(summaries) == 0:
    cu.error("Nerasta jokios tinkamos statistikos, kad būtų galima kurti grafiką!")
    exit()

# Sort summaries by
# 1) term start (year)
# 2) type (semester) (I -> II -> III)
summaries.sort(key=lambda s: (s.term_start, s.type_as_int))

# Get the last (the most recent) statistical data and cache student names for later use
student_cache = [s.name for s in summaries[-1].students]

# Go over each summary and use it to create graph points
_temp_subject_name_dict = {}
graph = StudentAveragesGraph([s.representable_name for s in summaries])
for i, summary in enumerate(summaries):
    cu.info(f"Nagrinėjamas {summary.period_name} ({summary.term_start}-{summary.term_end})")
    _cached_subs = []
    for student in summary.students:

        # If student name is not in cache, ignore them
        if student.name not in student_cache:
            cu.warn(f"Mokinys '{student.name}' ignoruojamas, nes nėra naujausioje suvestinėje")
            continue

        subjects = student.get_graphing_subjects()

        # Notify user regarding different generic and original subject names
        # Not in use as of now, but future proofing
        for subject in subjects:
            if subject.name not in _cached_subs:
                if subject.name != subject.generic_name:
                    cu.warn(f"Dalykas '{subject.name}' automatiškai pervadintas į '{subject.generic_name}'")
                _cached_subs.append(subject.name)

            if _temp_subject_name_dict.get(subject.name, None) is None:
                _temp_subject_name_dict[subject.name] = (subject.generic_name, summary.grade_name_as_int)

        graph.get_or_create_student(student.name)[i] = student.average

if DEBUG:
    for key in _temp_subject_name_dict.keys():
        name, grade = _temp_subject_name_dict[key]
        print(key, "->", name, f"({grade} kl.)")

graph.graph(
    summary.grade_name + " mokinių vidurkių pokytis",
    anonymize_names=True,
    use_styled_colouring=True, use_experimental_legend=True
)

"""
# This is for later when I find out the spec
inp = cu.input("Ar Jūs norite modulius kartu pateikti grafike (dažniausiai neverta)? ") or "ne"
cleaned_inp = inp.lower().strip()
if cleaned_inp == "taip" or cleaned_inp == "t":
    cu.info("Moduliai bus įtraukti į grafiką")
else:
    cu.info("Moduliai nebus įtraukti į grafiką")"""
